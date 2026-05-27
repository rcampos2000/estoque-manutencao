"""
app.py — Sistema de Controle de Peças de Reposição — Manutenção Industrial
Versão 1.0 | CustomTkinter + SQLite
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import os, sys, io
from datetime import datetime
import database as db

# ── Tema ──────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

AZUL       = "#1565C0"
AZUL_CLARO = "#1E88E5"
VERDE      = "#2E7D32"
VERMELHO   = "#C62828"
AMARELO    = "#F57F17"
CINZA_BG   = "#1a1a2e"
CINZA_CARD = "#16213e"
CINZA_ITEM = "#0f3460"
BRANCO     = "#FFFFFF"
TEXTO_SEC  = "#B0BEC5"

USUARIO_ATUAL = {"id": None, "nome": "", "perfil": ""}


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITÁRIOS
# ══════════════════════════════════════════════════════════════════════════════

def label_tipo(tipo):
    mapa = {"entrada":"✅ Entrada","saida":"📤 Saída","ajuste":"🔧 Ajuste","devolucao":"↩️ Devolução"}
    return mapa.get(tipo, tipo)


def formata_moeda(v):
    try: return f"R$ {float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    except: return "R$ 0,00"


def cor_status(qtd, minimo):
    if qtd == 0:      return "#ef5350"
    if qtd <= minimo: return "#FFA726"
    return "#66BB6A"


# ══════════════════════════════════════════════════════════════════════════════
#  TELA DE LOGIN
# ══════════════════════════════════════════════════════════════════════════════

class TelaLogin(ctk.CTkToplevel):
    def __init__(self, callback_ok):
        super().__init__()
        self.callback_ok = callback_ok
        self.title("Login — Controle de Estoque Manutenção")
        self.geometry("420x520")
        self.resizable(False, False)
        self.configure(fg_color=CINZA_BG)
        self.grab_set()
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="⚙️", font=ctk.CTkFont(size=60)).pack(pady=(40,5))
        ctk.CTkLabel(self, text="Controle de Estoque",
                     font=ctk.CTkFont(size=22, weight="bold")).pack()
        ctk.CTkLabel(self, text="Peças de Reposição — Manutenção",
                     font=ctk.CTkFont(size=13), text_color=TEXTO_SEC).pack(pady=(2,30))

        frame = ctk.CTkFrame(self, fg_color=CINZA_CARD, corner_radius=12)
        frame.pack(padx=40, fill="x")

        ctk.CTkLabel(frame, text="Usuário", anchor="w").pack(padx=20, pady=(20,2), fill="x")
        self.e_login = ctk.CTkEntry(frame, placeholder_text="login", height=40)
        self.e_login.pack(padx=20, fill="x")

        ctk.CTkLabel(frame, text="Senha", anchor="w").pack(padx=20, pady=(12,2), fill="x")
        self.e_senha = ctk.CTkEntry(frame, placeholder_text="senha", show="*", height=40)
        self.e_senha.pack(padx=20, fill="x")
        self.e_senha.bind("<Return>", lambda e: self._login())

        self.lbl_erro = ctk.CTkLabel(frame, text="", text_color="#ef5350", font=ctk.CTkFont(size=12))
        self.lbl_erro.pack(pady=(6,0))

        ctk.CTkButton(frame, text="Entrar", height=44, font=ctk.CTkFont(size=14, weight="bold"),
                      command=self._login).pack(padx=20, pady=20, fill="x")

        ctk.CTkLabel(self, text="Admin padrão: admin / admin123",
                     font=ctk.CTkFont(size=11), text_color=TEXTO_SEC).pack(pady=10)

        self.e_login.focus()

    def _login(self):
        login = self.e_login.get().strip()
        senha = self.e_senha.get()
        user = db.autenticar(login, senha)
        if user:
            USUARIO_ATUAL["id"]     = user["id"]
            USUARIO_ATUAL["nome"]   = user["nome"]
            USUARIO_ATUAL["perfil"] = user["perfil"]
            self.destroy()
            self.callback_ok()
        else:
            self.lbl_erro.configure(text="⚠️ Usuário ou senha inválidos")


# ══════════════════════════════════════════════════════════════════════════════
#  WIDGET: TABELA CUSTOMIZADA (Treeview estilizado)
# ══════════════════════════════════════════════════════════════════════════════

class Tabela(ttk.Frame):
    def __init__(self, master, colunas, larguras=None, altura=18, **kw):
        super().__init__(master, **kw)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.Treeview",
                         background="#16213e", foreground="white",
                         fieldbackground="#16213e", rowheight=28,
                         borderwidth=0, font=("Segoe UI", 10))
        style.configure("Dark.Treeview.Heading",
                         background="#1565C0", foreground="white",
                         font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("Dark.Treeview",
                  background=[("selected","#1E88E5")],
                  foreground=[("selected","white")])

        self.tree = ttk.Treeview(self, columns=colunas, show="headings",
                                 height=altura, style="Dark.Treeview")
        for i, col in enumerate(colunas):
            w = larguras[i] if larguras else 120
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=40)

        vsb = ttk.Scrollbar(self, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1); self.columnconfigure(0, weight=1)

    def limpar(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def inserir(self, valores, tag=""):
        self.tree.insert("", "end", values=valores, tags=(tag,))

    def configurar_tag(self, tag, bg=None, fg=None):
        self.tree.tag_configure(tag, background=bg or "", foreground=fg or "white")

    def selecionado_id(self):
        sel = self.tree.selection()
        if sel:
            return self.tree.item(sel[0])["values"]
        return None

    def bind_duplo(self, callback):
        self.tree.bind("<Double-1>", callback)


# ══════════════════════════════════════════════════════════════════════════════
#  TELA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Controle de Peças de Reposição — Manutenção Industrial")
        self.geometry("1300x780")
        self.minsize(1100, 660)
        self.configure(fg_color=CINZA_BG)
        self._pagina_atual = None
        self._frames = {}
        self._build_layout()

    def _build_layout(self):
        # ── Sidebar ──────────────────────────────────────────────────────────
        self.sidebar = ctk.CTkFrame(self, width=220, fg_color=CINZA_CARD, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        ctk.CTkLabel(self.sidebar, text="⚙️ Estoque\nManutenção",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     justify="center").pack(pady=20)

        ctk.CTkFrame(self.sidebar, height=1, fg_color="#333").pack(fill="x", padx=10)

        menus = [
            ("🏠  Dashboard",      "dashboard"),
            ("📦  Peças",          "pecas"),
            ("🔍  Busca Rápida",   "busca"),
            ("📤  Movimentações",  "movimentacoes"),
            ("📊  Relatórios",     "relatorios"),
            ("🚨  Alertas",        "alertas"),
            ("🏭  Equipamentos",   "equipamentos"),
            ("🤝  Fornecedores",   "fornecedores"),
            ("👤  Usuários",       "usuarios"),
        ]

        self._botoes_menu = {}
        for label, page in menus:
            btn = ctk.CTkButton(
                self.sidebar, text=label, anchor="w",
                font=ctk.CTkFont(size=13), height=42,
                fg_color="transparent", hover_color=CINZA_ITEM,
                corner_radius=8,
                command=lambda p=page: self.ir_para(p)
            )
            btn.pack(fill="x", padx=8, pady=2)
            self._botoes_menu[page] = btn

        ctk.CTkFrame(self.sidebar, height=1, fg_color="#333").pack(fill="x", padx=10, pady=10)

        self.lbl_usuario = ctk.CTkLabel(
            self.sidebar, text=f"👤 {USUARIO_ATUAL['nome']}\n({USUARIO_ATUAL['perfil']})",
            font=ctk.CTkFont(size=11), text_color=TEXTO_SEC, justify="center")
        self.lbl_usuario.pack(pady=4)

        ctk.CTkButton(self.sidebar, text="Sair", fg_color=VERMELHO, hover_color="#b71c1c",
                      height=36, command=self._sair).pack(padx=10, pady=10, fill="x")

        # ── Área de conteúdo ──────────────────────────────────────────────────
        self.area = ctk.CTkFrame(self, fg_color=CINZA_BG, corner_radius=0)
        self.area.pack(side="right", fill="both", expand=True)

        # Inicializar todas as páginas
        self._frames["dashboard"]     = PaginaDashboard(self.area, self)
        self._frames["pecas"]         = PaginaPecas(self.area, self)
        self._frames["busca"]         = PaginaBusca(self.area, self)
        self._frames["movimentacoes"] = PaginaMovimentacoes(self.area, self)
        self._frames["relatorios"]    = PaginaRelatorios(self.area, self)
        self._frames["alertas"]       = PaginaAlertas(self.area, self)
        self._frames["equipamentos"]  = PaginaEquipamentos(self.area, self)
        self._frames["fornecedores"]  = PaginaFornecedores(self.area, self)
        self._frames["usuarios"]      = PaginaUsuarios(self.area, self)

        self.ir_para("dashboard")

    def ir_para(self, pagina):
        if self._pagina_atual:
            self._frames[self._pagina_atual].pack_forget()
            btn = self._botoes_menu.get(self._pagina_atual)
            if btn: btn.configure(fg_color="transparent")

        self._pagina_atual = pagina
        frame = self._frames[pagina]
        frame.pack(fill="both", expand=True)
        btn = self._botoes_menu.get(pagina)
        if btn: btn.configure(fg_color=AZUL)

        if hasattr(frame, "atualizar"):
            frame.atualizar()

    def _sair(self):
        if messagebox.askyesno("Sair", "Deseja sair do sistema?"):
            self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
#  PÁGINA: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

class PaginaDashboard(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=CINZA_BG, corner_radius=0)
        self.app = app
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="📊 Dashboard", font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(15,5), padx=20, anchor="w")
        ctk.CTkLabel(self, text=f"Bem-vindo, {USUARIO_ATUAL['nome']}!", text_color=TEXTO_SEC).pack(padx=20, anchor="w")

        # KPI Cards
        self.frame_kpis = ctk.CTkFrame(self, fg_color="transparent")
        self.frame_kpis.pack(fill="x", padx=20, pady=15)

        # Tabelas lado a lado
        row_frames = ctk.CTkFrame(self, fg_color="transparent")
        row_frames.pack(fill="both", expand=True, padx=20, pady=(0,15))

        # Top consumo
        f_top = ctk.CTkFrame(row_frames, fg_color=CINZA_CARD, corner_radius=10)
        f_top.pack(side="left", fill="both", expand=True, padx=(0,8))
        ctk.CTkLabel(f_top, text="🔥 Top Consumo (30 dias)", font=ctk.CTkFont(weight="bold")).pack(pady=10)
        self.tab_top = Tabela(f_top, ["Peça","Código","Saídas"],
                              larguras=[220,90,70], altura=10)
        self.tab_top.pack(fill="both", expand=True, padx=10, pady=(0,10))

        # Alertas
        f_alerta = ctk.CTkFrame(row_frames, fg_color=CINZA_CARD, corner_radius=10)
        f_alerta.pack(side="left", fill="both", expand=True)
        ctk.CTkLabel(f_alerta, text="🚨 Alertas de Estoque Crítico", font=ctk.CTkFont(weight="bold")).pack(pady=10)
        self.tab_alertas = Tabela(f_alerta, ["Peça","Qtd. Atual","Status"],
                                  larguras=[200,90,90], altura=10)
        self.tab_alertas.pack(fill="both", expand=True, padx=10, pady=(0,10))

    def _card_kpi(self, parent, titulo, valor, cor, icone):
        card = ctk.CTkFrame(parent, fg_color=CINZA_CARD, corner_radius=12)
        card.pack(side="left", fill="both", expand=True, padx=4)
        ctk.CTkLabel(card, text=icone, font=ctk.CTkFont(size=28)).pack(pady=(15,2))
        ctk.CTkLabel(card, text=str(valor), font=ctk.CTkFont(size=26, weight="bold"),
                     text_color=cor).pack()
        ctk.CTkLabel(card, text=titulo, font=ctk.CTkFont(size=11),
                     text_color=TEXTO_SEC).pack(pady=(2,15))

    def atualizar(self):
        kpis = db.get_kpis()
        for w in self.frame_kpis.winfo_children():
            w.destroy()
        self._card_kpi(self.frame_kpis, "Total de Peças",     kpis["total_pecas"],   BRANCO,    "📦")
        self._card_kpi(self.frame_kpis, "Estoque Crítico",    kpis["criticos"],      AMARELO,   "⚠️")
        self._card_kpi(self.frame_kpis, "Zerados",            kpis["zerados"],       VERMELHO,  "🔴")
        self._card_kpi(self.frame_kpis, "Valor em Estoque",   formata_moeda(kpis["valor_total"]), VERDE, "💰")
        self._card_kpi(self.frame_kpis, "Entradas (30d)",     kpis["entradas_30d"],  AZUL_CLARO,"✅")
        self._card_kpi(self.frame_kpis, "Saídas (30d)",       kpis["saidas_30d"],    "#FF7043",  "📤")
        self._card_kpi(self.frame_kpis, "Alertas Ativos",     kpis["alertas"],       AMARELO,   "🔔")

        self.tab_top.limpar()
        for r in db.get_top_consumo():
            self.tab_top.inserir([r["nome"], r["codigo"], f"{r['total_saido']:.1f}"])

        self.tab_alertas.limpar()
        for r in db.get_alertas_nao_lidos()[:15]:
            status = "🔴 ZERADO" if r["quantidade"]==0 else "⚠️ CRÍTICO"
            tag = "zero" if r["quantidade"]==0 else "critico"
            self.tab_alertas.inserir([r["peca_nome"], f"{r['quantidade']:.1f}", status], tag)
        self.tab_alertas.configurar_tag("zero",    bg="#5D1A1A")
        self.tab_alertas.configurar_tag("critico", bg="#5D4000")


# ══════════════════════════════════════════════════════════════════════════════
#  PÁGINA: PEÇAS
# ══════════════════════════════════════════════════════════════════════════════

class PaginaPecas(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=CINZA_BG, corner_radius=0)
        self.app = app
        self._build()

    def _build(self):
        # Cabeçalho
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(15,5))
        ctk.CTkLabel(hdr, text="📦 Peças de Reposição",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(side="left")
        ctk.CTkButton(hdr, text="＋ Nova Peça", fg_color=VERDE, hover_color="#1B5E20",
                      width=130, command=self._nova_peca).pack(side="right")
        ctk.CTkButton(hdr, text="📤 Exportar Excel", fg_color=AZUL, width=140,
                      command=self._exportar).pack(side="right", padx=8)

        # Filtros
        filtros = ctk.CTkFrame(self, fg_color=CINZA_CARD, corner_radius=10)
        filtros.pack(fill="x", padx=20, pady=5)

        ctk.CTkLabel(filtros, text="🔍 Pesquisar:").pack(side="left", padx=(15,5), pady=10)
        self.e_busca = ctk.CTkEntry(filtros, placeholder_text="Nome, código ou código de barras...",
                                    width=280, height=36)
        self.e_busca.pack(side="left", padx=5)
        self.e_busca.bind("<Return>", lambda e: self.atualizar())

        ctk.CTkLabel(filtros, text="Categoria:").pack(side="left", padx=(15,5))
        self.cats = db.listar_categorias()
        cat_names = ["Todas"] + [c["nome"] for c in self.cats]
        self.cmb_cat = ctk.CTkComboBox(filtros, values=cat_names, width=150, height=36,
                                       command=lambda v: self.atualizar())
        self.cmb_cat.pack(side="left", padx=5)

        self.var_critico = ctk.CTkCheckBox(filtros, text="Apenas críticos",
                                           command=self.atualizar)
        self.var_critico.pack(side="left", padx=15)

        ctk.CTkButton(filtros, text="🔍 Buscar", width=90, height=36,
                      command=self.atualizar).pack(side="left", padx=5)
        ctk.CTkButton(filtros, text="✖ Limpar", width=80, height=36,
                      fg_color="gray40", command=self._limpar_filtro).pack(side="left", padx=2)

        # Tabela
        self.tabela = Tabela(
            self,
            colunas=["Código","Cód.Barras","Nome","Categoria","Qtd","Mín","Unid","Localização","Custo Unit.","Valor Total"],
            larguras=[90, 110, 220, 110, 60, 55, 50, 120, 95, 95],
            altura=22
        )
        self.tabela.pack(fill="both", expand=True, padx=20, pady=(5,5))
        self.tabela.bind_duplo(self._editar_selecionado)

        self.lbl_total = ctk.CTkLabel(self, text="", text_color=TEXTO_SEC, font=ctk.CTkFont(size=11))
        self.lbl_total.pack(pady=5, anchor="e", padx=20)

        # Ações
        acoes = ctk.CTkFrame(self, fg_color="transparent")
        acoes.pack(fill="x", padx=20, pady=(0,10))
        ctk.CTkButton(acoes, text="✏️ Editar", width=110, command=self._editar_selecionado).pack(side="left", padx=4)
        ctk.CTkButton(acoes, text="📤 Registrar Saída", width=140, fg_color="#E65100",
                      command=lambda: self._movimentar("saida")).pack(side="left", padx=4)
        ctk.CTkButton(acoes, text="📥 Registrar Entrada", width=145, fg_color=VERDE,
                      command=lambda: self._movimentar("entrada")).pack(side="left", padx=4)
        ctk.CTkButton(acoes, text="🏷️ Gerar Etiqueta", width=130, fg_color="#7B1FA2",
                      command=self._gerar_etiqueta).pack(side="left", padx=4)
        ctk.CTkButton(acoes, text="🗑️ Excluir", width=95, fg_color=VERMELHO,
                      command=self._excluir).pack(side="right", padx=4)

    def _limpar_filtro(self):
        self.e_busca.delete(0, "end")
        self.cmb_cat.set("Todas")
        self.var_critico.deselect()
        self.atualizar()

    def atualizar(self):
        filtro = self.e_busca.get().strip()
        cat_nome = self.cmb_cat.get()
        cat_id = None
        if cat_nome != "Todas":
            for c in self.cats:
                if c["nome"] == cat_nome:
                    cat_id = c["id"]; break
        critico = self.var_critico.get()

        rows = db.listar_pecas(filtro=filtro, categoria_id=cat_id, apenas_criticos=critico)
        self.tabela.limpar()
        valor_tot = 0
        for r in rows:
            qtd   = r["quantidade"] or 0
            emin  = r["estoque_minimo"] or 0
            vtot  = qtd * (r["custo_unitario"] or 0)
            valor_tot += vtot

            if qtd == 0:       tag = "zero"
            elif qtd <= emin:  tag = "critico"
            else:              tag = ""

            self.tabela.inserir([
                r["codigo"], r["codigo_barras"] or "", r["nome"],
                r["categoria_nome"] or "", f"{qtd:.1f}", f"{emin:.1f}",
                r["unidade"], r["localizacao_nome"] or "",
                formata_moeda(r["custo_unitario"]), formata_moeda(vtot)
            ], tag)

        self.tabela.configurar_tag("zero",    bg="#5D1A1A")
        self.tabela.configurar_tag("critico", bg="#5D4000")
        self.lbl_total.configure(
            text=f"Total: {len(rows)} peças  |  Valor em estoque: {formata_moeda(valor_tot)}")
        self.cats = db.listar_categorias()

    def _peca_selecionada(self):
        vals = self.tabela.selecionado_id()
        if not vals:
            messagebox.showwarning("Atenção", "Selecione uma peça na lista.")
            return None
        codigo = vals[0]
        pecas = db.listar_pecas(filtro=codigo)
        for p in pecas:
            if p["codigo"] == codigo:
                return p
        return None

    def _nova_peca(self):
        FormularioPeca(self, self.app, None, callback=self.atualizar)

    def _editar_selecionado(self, event=None):
        p = self._peca_selecionada()
        if p: FormularioPeca(self, self.app, p, callback=self.atualizar)

    def _excluir(self):
        p = self._peca_selecionada()
        if not p: return
        if messagebox.askyesno("Confirmar", f"Excluir '{p['nome']}'?"):
            db.excluir_peca(p["id"])
            self.atualizar()

    def _movimentar(self, tipo):
        p = self._peca_selecionada()
        if p: DialogMovimentacao(self, self.app, p, tipo, callback=self.atualizar)

    def _gerar_etiqueta(self):
        p = self._peca_selecionada()
        if p: gerar_etiqueta(p)

    def _exportar(self):
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Salvar relatório de estoque",
            initialfile=f"estoque_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if caminho:
            db.exportar_estoque_excel(caminho)
            messagebox.showinfo("Sucesso", f"Exportado:\n{caminho}")


# ══════════════════════════════════════════════════════════════════════════════
#  FORMULÁRIO: CADASTRO / EDIÇÃO DE PEÇA
# ══════════════════════════════════════════════════════════════════════════════

class FormularioPeca(ctk.CTkToplevel):
    def __init__(self, master, app, peca, callback):
        super().__init__(master)
        self.app = app
        self.peca = peca
        self.callback = callback
        self.title("Nova Peça" if not peca else "Editar Peça")
        self.geometry("620x720")
        self.resizable(False, False)
        self.configure(fg_color=CINZA_BG)
        self.grab_set()
        self._cats = db.listar_categorias()
        self._forns = db.listar_fornecedores()
        self._locs  = db.listar_localizacoes()
        self._build()
        if peca: self._preencher()

    def _campo(self, parent, label, **kw):
        ctk.CTkLabel(parent, text=label, anchor="w").pack(fill="x", padx=5, pady=(8,1))
        e = ctk.CTkEntry(parent, height=36, **kw)
        e.pack(fill="x", padx=5)
        return e

    def _combo(self, parent, label, valores, width=None):
        ctk.CTkLabel(parent, text=label, anchor="w").pack(fill="x", padx=5, pady=(8,1))
        c = ctk.CTkComboBox(parent, values=valores, height=36)
        c.pack(fill="x", padx=5)
        return c

    def _build(self):
        # Scroll
        scroll = ctk.CTkScrollableFrame(self, fg_color=CINZA_BG)
        scroll.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(scroll, text="Nova Peça" if not self.peca else "Editar Peça",
                     font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(0,10))

        # Linha 1: Código / Código de barras
        row1 = ctk.CTkFrame(scroll, fg_color="transparent")
        row1.pack(fill="x")
        f1 = ctk.CTkFrame(row1, fg_color="transparent"); f1.pack(side="left", fill="x", expand=True, padx=(0,5))
        f2 = ctk.CTkFrame(row1, fg_color="transparent"); f2.pack(side="left", fill="x", expand=True)
        self.e_codigo   = self._campo(f1, "Código *")
        self.e_barcode  = self._campo(f2, "Código de Barras")

        # Nome
        self.e_nome = self._campo(scroll, "Nome *", placeholder_text="Nome da peça")

        # Descrição
        ctk.CTkLabel(scroll, text="Descrição", anchor="w").pack(fill="x", padx=5, pady=(8,1))
        self.e_desc = ctk.CTkTextbox(scroll, height=60, fg_color=CINZA_ITEM)
        self.e_desc.pack(fill="x", padx=5)

        # Categoria / Fornecedor / Localização
        row2 = ctk.CTkFrame(scroll, fg_color="transparent")
        row2.pack(fill="x")
        for label, attr, lista, key in [
            ("Categoria", "_cmb_cat",  [""] + [c["nome"] for c in self._cats],  None),
            ("Fornecedor","_cmb_forn", [""] + [f["nome"] for f in self._forns], None),
            ("Localização","_cmb_loc", [""] + [l["nome"] for l in self._locs],  None),
        ]:
            fr = ctk.CTkFrame(row2, fg_color="transparent")
            fr.pack(side="left", fill="x", expand=True, padx=3)
            cmb = self._combo(fr, label, lista)
            setattr(self, attr, cmb)

        # Quantidade / Mínimo / Máximo / Unidade
        row3 = ctk.CTkFrame(scroll, fg_color="transparent")
        row3.pack(fill="x")
        for label, attr, defval in [
            ("Quantidade Inicial","_e_qtd","0"),
            ("Estoque Mínimo","_e_emin","1"),
            ("Estoque Máximo","_e_emax","100"),
            ("Unidade","_e_unid","UN"),
        ]:
            fr = ctk.CTkFrame(row3, fg_color="transparent")
            fr.pack(side="left", fill="x", expand=True, padx=3)
            e = self._campo(fr, label)
            e.insert(0, defval)
            setattr(self, attr, e)

        # Custo / Preço venda
        row4 = ctk.CTkFrame(scroll, fg_color="transparent")
        row4.pack(fill="x")
        for label, attr in [("Custo Unitário (R$)","_e_custo"),("Preço de Venda (R$)","_e_pvenda")]:
            fr = ctk.CTkFrame(row4, fg_color="transparent")
            fr.pack(side="left", fill="x", expand=True, padx=3)
            e = self._campo(fr, label)
            e.insert(0, "0.00")
            setattr(self, attr, e)

        # Botões
        btns = ctk.CTkFrame(scroll, fg_color="transparent")
        btns.pack(fill="x", pady=15)
        ctk.CTkButton(btns, text="💾 Salvar", fg_color=VERDE, command=self._salvar).pack(side="left", padx=10, expand=True, fill="x")
        ctk.CTkButton(btns, text="✖ Cancelar", fg_color="gray40", command=self.destroy).pack(side="left", padx=10, expand=True, fill="x")

    def _preencher(self):
        p = self.peca
        self.e_codigo.insert(0, p["codigo"] or "")
        self.e_barcode.insert(0, p["codigo_barras"] or "")
        self.e_nome.insert(0, p["nome"] or "")
        self.e_desc.insert("0.0", p["descricao"] or "")

        for lista, attr, nome_col in [
            (self._cats, "_cmb_cat",  "categoria_nome"),
            (self._forns,"_cmb_forn", "fornecedor_nome"),
            (self._locs, "_cmb_loc",  "localizacao_nome"),
        ]:
            val = p[nome_col] or ""
            getattr(self, attr).set(val)

        self._e_qtd.delete(0,"end");   self._e_qtd.insert(0,  str(p["quantidade"] or 0))
        self._e_emin.delete(0,"end");  self._e_emin.insert(0, str(p["estoque_minimo"] or 1))
        self._e_emax.delete(0,"end");  self._e_emax.insert(0, str(p["estoque_maximo"] or 100))
        self._e_unid.delete(0,"end");  self._e_unid.insert(0, p["unidade"] or "UN")
        self._e_custo.delete(0,"end"); self._e_custo.insert(0, str(p["custo_unitario"] or 0))
        self._e_pvenda.delete(0,"end");self._e_pvenda.insert(0,str(p["preco_venda"] or 0))

    def _id_por_nome(self, lista, nome, campo="nome"):
        for item in lista:
            if item[campo] == nome:
                return item["id"]
        return None

    def _salvar(self):
        codigo = self.e_codigo.get().strip()
        nome   = self.e_nome.get().strip()
        if not codigo or not nome:
            messagebox.showerror("Erro", "Código e Nome são obrigatórios.")
            return
        try:
            dados = {
                "codigo":        codigo,
                "codigo_barras": self.e_barcode.get().strip() or None,
                "nome":          nome,
                "descricao":     self.e_desc.get("0.0","end").strip(),
                "categoria_id":  self._id_por_nome(self._cats,  self._cmb_cat.get()),
                "fornecedor_id": self._id_por_nome(self._forns, self._cmb_forn.get()),
                "localizacao_id":self._id_por_nome(self._locs,  self._cmb_loc.get()),
                "unidade":       self._e_unid.get().strip() or "UN",
                "quantidade":    float(self._e_qtd.get()   or 0),
                "estoque_minimo":float(self._e_emin.get()  or 1),
                "estoque_maximo":float(self._e_emax.get()  or 100),
                "custo_unitario":float(self._e_custo.get() or 0),
                "preco_venda":   float(self._e_pvenda.get()or 0),
            }
        except ValueError:
            messagebox.showerror("Erro", "Valores numéricos inválidos.")
            return

        if self.peca:
            db.atualizar_peca(self.peca["id"], dados)
        else:
            db.inserir_peca(dados)

        self.callback()
        self.destroy()
        messagebox.showinfo("Sucesso", "Peça salva com sucesso!")


# ══════════════════════════════════════════════════════════════════════════════
#  DIÁLOGO: MOVIMENTAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

class DialogMovimentacao(ctk.CTkToplevel):
    def __init__(self, master, app, peca, tipo, callback):
        super().__init__(master)
        self.app = app
        self.peca = peca
        self.tipo_mov = tipo
        self.callback = callback
        titulos = {"saida":"📤 Registrar Saída","entrada":"📥 Registrar Entrada",
                   "ajuste":"🔧 Ajuste de Estoque","devolucao":"↩️ Devolução"}
        self.title(titulos.get(tipo, "Movimentação"))
        self.geometry("480x520")
        self.resizable(False, False)
        self.configure(fg_color=CINZA_BG)
        self.grab_set()
        self._equips = db.listar_equipamentos()
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text=f"Peça: {self.peca['nome']}",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(20,2), padx=20, anchor="w")
        ctk.CTkLabel(self, text=f"Código: {self.peca['codigo']}  |  Qtd atual: {self.peca['quantidade']:.1f} {self.peca['unidade']}",
                     text_color=TEXTO_SEC).pack(padx=20, anchor="w")

        frame = ctk.CTkScrollableFrame(self, fg_color=CINZA_CARD, corner_radius=12)
        frame.pack(fill="both", expand=True, padx=20, pady=15)

        ctk.CTkLabel(frame, text="Quantidade *", anchor="w").pack(fill="x", padx=10, pady=(10,2))
        self.e_qtd = ctk.CTkEntry(frame, height=40, placeholder_text="0")
        self.e_qtd.pack(fill="x", padx=10)
        self.e_qtd.focus()

        ctk.CTkLabel(frame, text="Tipo de Movimentação", anchor="w").pack(fill="x", padx=10, pady=(10,2))
        self.cmb_tipo = ctk.CTkComboBox(frame, height=36,
            values=["saida","entrada","ajuste","devolucao"])
        self.cmb_tipo.set(self.tipo_mov)
        self.cmb_tipo.pack(fill="x", padx=10)

        ctk.CTkLabel(frame, text="Equipamento / TAG", anchor="w").pack(fill="x", padx=10, pady=(10,2))
        equip_nomes = [""] + [f"{e['tag']} — {e['nome']}" for e in self._equips]
        self.cmb_equip = ctk.CTkComboBox(frame, values=equip_nomes, height=36)
        self.cmb_equip.pack(fill="x", padx=10)

        ctk.CTkLabel(frame, text="Nº da O.S.", anchor="w").pack(fill="x", padx=10, pady=(10,2))
        self.e_os = ctk.CTkEntry(frame, height=36, placeholder_text="OS-0001")
        self.e_os.pack(fill="x", padx=10)

        ctk.CTkLabel(frame, text="Motivo / Descrição", anchor="w").pack(fill="x", padx=10, pady=(10,2))
        self.e_motivo = ctk.CTkEntry(frame, height=36, placeholder_text="Manutenção preventiva...")
        self.e_motivo.pack(fill="x", padx=10)

        ctk.CTkLabel(frame, text="Observação", anchor="w").pack(fill="x", padx=10, pady=(10,2))
        self.e_obs = ctk.CTkTextbox(frame, height=60, fg_color=CINZA_ITEM)
        self.e_obs.pack(fill="x", padx=10, pady=(0,10))

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=20, pady=(0,15))
        cor = VERDE if self.tipo_mov=="entrada" else VERMELHO if self.tipo_mov=="saida" else AZUL
        ctk.CTkButton(btns, text="✔ Confirmar", fg_color=cor,
                      command=self._confirmar).pack(side="left", fill="x", expand=True, padx=5)
        ctk.CTkButton(btns, text="✖ Cancelar", fg_color="gray40",
                      command=self.destroy).pack(side="left", fill="x", expand=True, padx=5)

    def _confirmar(self):
        try:
            qtd = float(self.e_qtd.get() or 0)
            if qtd <= 0: raise ValueError
        except ValueError:
            messagebox.showerror("Erro", "Quantidade inválida.")
            return

        tipo = self.cmb_tipo.get()
        if tipo == "saida" and qtd > self.peca["quantidade"]:
            if not messagebox.askyesno("Atenção",
               f"Estoque insuficiente!\nAtual: {self.peca['quantidade']:.1f}\nDeseja continuar mesmo assim?"):
                return

        equip_id = None
        sel = self.cmb_equip.get()
        if sel:
            tag = sel.split(" — ")[0]
            for e in self._equips:
                if e["tag"] == tag:
                    equip_id = e["id"]; break

        db.registrar_movimentacao(
            peca_id=self.peca["id"], tipo=tipo, quantidade=qtd,
            usuario_id=USUARIO_ATUAL["id"], custo=self.peca["custo_unitario"],
            equip_id=equip_id, os_num=self.e_os.get().strip(),
            motivo=self.e_motivo.get().strip(), obs=self.e_obs.get("0.0","end").strip()
        )
        self.callback()
        self.destroy()
        messagebox.showinfo("Sucesso", f"Movimentação registrada!\nNova quantidade: {qtd}")


# ══════════════════════════════════════════════════════════════════════════════
#  PÁGINA: BUSCA RÁPIDA (com suporte a scanner)
# ══════════════════════════════════════════════════════════════════════════════

class PaginaBusca(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=CINZA_BG, corner_radius=0)
        self.app = app
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="🔍 Busca Rápida",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(15,5), padx=20, anchor="w")
        ctk.CTkLabel(self, text="Pesquise por código de barras (leitora USB ou teclado), código ou nome",
                     text_color=TEXTO_SEC).pack(padx=20, anchor="w")

        # Barra de busca grande
        frame_busca = ctk.CTkFrame(self, fg_color=CINZA_CARD, corner_radius=12)
        frame_busca.pack(fill="x", padx=20, pady=15)

        ctk.CTkLabel(frame_busca, text="📷 Código de Barras / Pesquisa:",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(15,5))

        row = ctk.CTkFrame(frame_busca, fg_color="transparent")
        row.pack(padx=20, pady=(0,15), fill="x")
        self.e_scan = ctk.CTkEntry(row, height=52, font=ctk.CTkFont(size=18),
                                   placeholder_text="🔍  Aponte a leitora ou digite aqui...")
        self.e_scan.pack(side="left", fill="x", expand=True, padx=(0,8))
        self.e_scan.bind("<Return>", self._buscar)
        ctk.CTkButton(row, text="🔍 BUSCAR", height=52, width=110,
                      font=ctk.CTkFont(size=14, weight="bold"),
                      command=self._buscar).pack(side="left")

        self.e_scan.focus()

        # Resultado
        self.frame_resultado = ctk.CTkScrollableFrame(self, fg_color=CINZA_CARD, corner_radius=12, height=260)
        self.frame_resultado.pack(fill="x", padx=20, pady=(0,10))
        self._lbl_resultado = ctk.CTkLabel(self.frame_resultado,
                                            text="🔎 Digite um código ou use a leitora de barras",
                                            text_color=TEXTO_SEC, font=ctk.CTkFont(size=14))
        self._lbl_resultado.pack(pady=40)

        # Ações rápidas
        self.frame_acoes = ctk.CTkFrame(self, fg_color="transparent")
        self.frame_acoes.pack(fill="x", padx=20, pady=5)
        self._peca_encontrada = None

    def _buscar(self, event=None):
        termo = self.e_scan.get().strip()
        if not termo: return

        for w in self.frame_resultado.winfo_children(): w.destroy()
        for w in self.frame_acoes.winfo_children(): w.destroy()

        peca = db.buscar_peca_barcode(termo)
        if not peca:
            # Tenta busca por nome
            rows = db.listar_pecas(filtro=termo)
            if rows:
                peca = rows[0]

        if not peca:
            ctk.CTkLabel(self.frame_resultado,
                         text=f"❌ Nenhuma peça encontrada para: '{termo}'",
                         text_color=VERMELHO, font=ctk.CTkFont(size=14)).pack(pady=30)
            return

        self._peca_encontrada = peca
        self._mostrar_peca(peca)
        self.e_scan.delete(0, "end")

    def _mostrar_peca(self, p):
        qtd = p["quantidade"] or 0
        emin = p["estoque_minimo"] or 0
        cor = cor_status(qtd, emin)

        ctk.CTkLabel(self.frame_resultado, text=f"✅ Peça Encontrada",
                     text_color=VERDE, font=ctk.CTkFont(size=15, weight="bold")).pack(pady=(15,5))

        grid = ctk.CTkFrame(self.frame_resultado, fg_color="transparent")
        grid.pack(fill="x", padx=20, pady=5)

        campos = [
            ("Código:",       p["codigo"]),
            ("Código Barras:",p["codigo_barras"] or "—"),
            ("Nome:",         p["nome"]),
            ("Categoria:",    p["categoria_nome"] or "—"),
            ("Localização:",  p["localizacao_nome"] or "—"),
            ("Fornecedor:",   p["fornecedor_nome"] or "—"),
            ("Custo Unit.:",  formata_moeda(p["custo_unitario"])),
        ]

        for i, (lbl, val) in enumerate(campos):
            ctk.CTkLabel(grid, text=lbl, text_color=TEXTO_SEC, anchor="e", width=130).grid(
                row=i, column=0, sticky="e", padx=(0,8), pady=3)
            ctk.CTkLabel(grid, text=str(val), anchor="w").grid(
                row=i, column=1, sticky="w", pady=3)

        # Quantidade grande
        ctk.CTkLabel(self.frame_resultado,
                     text=f"{qtd:.1f} {p['unidade']}",
                     font=ctk.CTkFont(size=36, weight="bold"),
                     text_color=cor).pack(pady=5)
        ctk.CTkLabel(self.frame_resultado,
                     text=f"Estoque mínimo: {emin:.1f}  |  Máximo: {p['estoque_maximo']:.1f}",
                     text_color=TEXTO_SEC).pack()

        # Ações rápidas
        ctk.CTkButton(self.frame_acoes, text="📤 Saída Rápida", fg_color=VERMELHO,
                      command=lambda: DialogMovimentacao(self, self.app, p, "saida",
                                                         callback=lambda: None)
                      ).pack(side="left", padx=5)
        ctk.CTkButton(self.frame_acoes, text="📥 Entrada Rápida", fg_color=VERDE,
                      command=lambda: DialogMovimentacao(self, self.app, p, "entrada",
                                                         callback=lambda: None)
                      ).pack(side="left", padx=5)
        ctk.CTkButton(self.frame_acoes, text="🏷️ Gerar Etiqueta", fg_color="#7B1FA2",
                      command=lambda: gerar_etiqueta(p)).pack(side="left", padx=5)
        ctk.CTkButton(self.frame_acoes, text="📋 Histórico", fg_color=AZUL,
                      command=lambda: self._ver_historico(p)).pack(side="left", padx=5)

    def _ver_historico(self, peca):
        HistoricoDialog(self, peca)

    def atualizar(self):
        self.e_scan.focus()


# ══════════════════════════════════════════════════════════════════════════════
#  PÁGINA: MOVIMENTAÇÕES
# ══════════════════════════════════════════════════════════════════════════════

class PaginaMovimentacoes(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=CINZA_BG, corner_radius=0)
        self.app = app
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="📋 Histórico de Movimentações",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(15,5), padx=20, anchor="w")

        filtros = ctk.CTkFrame(self, fg_color=CINZA_CARD, corner_radius=10)
        filtros.pack(fill="x", padx=20, pady=5)

        ctk.CTkLabel(filtros, text="Tipo:").pack(side="left", padx=(15,5), pady=10)
        self.cmb_tipo = ctk.CTkComboBox(filtros, width=130, height=36,
                                        values=["Todos","entrada","saida","ajuste","devolucao"])
        self.cmb_tipo.pack(side="left", padx=5)

        ctk.CTkLabel(filtros, text="De:").pack(side="left", padx=(10,5))
        self.e_data_ini = ctk.CTkEntry(filtros, width=110, height=36, placeholder_text="AAAA-MM-DD")
        self.e_data_ini.pack(side="left", padx=5)

        ctk.CTkLabel(filtros, text="Até:").pack(side="left", padx=(5,5))
        self.e_data_fim = ctk.CTkEntry(filtros, width=110, height=36, placeholder_text="AAAA-MM-DD")
        self.e_data_fim.pack(side="left", padx=5)

        ctk.CTkLabel(filtros, text="Peça:").pack(side="left", padx=(10,5))
        self.e_peca = ctk.CTkEntry(filtros, width=180, height=36, placeholder_text="Pesquisar peça...")
        self.e_peca.pack(side="left", padx=5)

        ctk.CTkButton(filtros, text="🔍 Filtrar", width=90, height=36,
                      command=self.atualizar).pack(side="left", padx=8)
        ctk.CTkButton(filtros, text="📤 Excel", width=80, height=36, fg_color=VERDE,
                      command=self._exportar).pack(side="left", padx=4)

        self.tabela = Tabela(
            self,
            colunas=["Data/Hora","Tipo","Peça","Código","Qtd","Ant.","Pos.","O.S.","Equipamento","Usuário"],
            larguras=[140,100,200,90,60,60,60,90,140,100],
            altura=22
        )
        self.tabela.pack(fill="both", expand=True, padx=20, pady=(5,15))

        self.lbl_total = ctk.CTkLabel(self, text="", text_color=TEXTO_SEC)
        self.lbl_total.pack(pady=5, anchor="e", padx=20)

    def atualizar(self):
        tipo = self.cmb_tipo.get()
        if tipo == "Todos": tipo = None
        ini = self.e_data_ini.get().strip() or None
        fim = self.e_data_fim.get().strip() or None

        rows = db.listar_movimentacoes(tipo=tipo, data_inicio=ini, data_fim=fim)
        filtro_peca = self.e_peca.get().strip().lower()

        self.tabela.limpar()
        count = 0
        for r in rows:
            nome = r["peca_nome"] or ""
            if filtro_peca and filtro_peca not in nome.lower():
                continue
            tag = r["tipo"]
            data = r["data_hora"][:16] if r["data_hora"] else ""
            self.tabela.inserir([
                data, label_tipo(r["tipo"]), nome,
                r["peca_codigo"] or "", f"{r['quantidade']:.1f}",
                f"{r['quantidade_ant']:.1f}", f"{r['quantidade_pos']:.1f}",
                r["os_numero"] or "", r["equipamento_tag"] or "",
                r["usuario_nome"] or ""
            ], tag)
            count += 1

        self.tabela.configurar_tag("entrada",   bg="#1B5E20")
        self.tabela.configurar_tag("saida",     bg="#7f0000")
        self.tabela.configurar_tag("ajuste",    bg="#33691E")
        self.tabela.configurar_tag("devolucao", bg="#1A237E")
        self.lbl_total.configure(text=f"Total: {count} movimentações")

    def _exportar(self):
        rows = db.listar_movimentacoes(limit=10000)
        if not rows: return
        import openpyxl
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"movimentacoes_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not caminho: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Movimentações"
        cols = ["Data/Hora","Tipo","Peça","Código","Quantidade","Qtd Ant.","Qtd Pos.","O.S.","Equipamento","Usuário","Motivo"]
        ws.append(cols)
        for r in rows:
            ws.append([r["data_hora"], r["tipo"], r["peca_nome"], r["peca_codigo"],
                       r["quantidade"], r["quantidade_ant"], r["quantidade_pos"],
                       r["os_numero"], r["equipamento_tag"], r["usuario_nome"], r["motivo"]])
        wb.save(caminho)
        messagebox.showinfo("Sucesso", f"Exportado: {caminho}")


# ══════════════════════════════════════════════════════════════════════════════
#  PÁGINA: RELATÓRIOS
# ══════════════════════════════════════════════════════════════════════════════

class PaginaRelatorios(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=CINZA_BG, corner_radius=0)
        self.app = app
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="📊 Relatórios",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(15,5), padx=20, anchor="w")

        frame = ctk.CTkScrollableFrame(self, fg_color=CINZA_BG)
        frame.pack(fill="both", expand=True, padx=20, pady=10)

        relatorios = [
            ("📦 Estoque Atual (Excel)", "Exporta todo o estoque atual com formatação profissional, alertas de itens críticos e valor total", self._rel_estoque_excel, AZUL),
            ("📤 Movimentações (Excel)", "Histórico completo de entradas e saídas em planilha Excel", self._rel_mov_excel, VERDE),
            ("🚨 Itens Críticos (Excel)", "Lista apenas peças abaixo do estoque mínimo", self._rel_criticos, AMARELO),
            ("📊 Consumo Mensal (Gráfico)", "Visualização de entradas e saídas nos últimos 12 meses", self._grafico_consumo, "#7B1FA2"),
            ("🔥 Top 10 Mais Consumidas", "Peças com maior volume de saídas nos últimos 30 dias", self._grafico_top, "#E65100"),
        ]

        for titulo, desc, cmd, cor in relatorios:
            card = ctk.CTkFrame(frame, fg_color=CINZA_CARD, corner_radius=10)
            card.pack(fill="x", pady=6)
            row = ctk.CTkFrame(card, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=15)
            ctk.CTkLabel(row, text=titulo, font=ctk.CTkFont(size=14, weight="bold")).pack(side="left", anchor="w")
            ctk.CTkButton(row, text="▶ Gerar", fg_color=cor, width=100, command=cmd).pack(side="right")
            ctk.CTkLabel(card, text=desc, text_color=TEXTO_SEC, anchor="w").pack(padx=20, pady=(0,12), anchor="w")

    def _rel_estoque_excel(self):
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"estoque_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if caminho:
            db.exportar_estoque_excel(caminho)
            messagebox.showinfo("Sucesso", f"Relatório gerado:\n{caminho}")

    def _rel_mov_excel(self):
        import openpyxl
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"movimentacoes_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not caminho: return
        rows = db.listar_movimentacoes(limit=50000)
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Movimentações"
        ws.append(["Data/Hora","Tipo","Peça","Código","Qtd","Qtd Ant","Qtd Pos","O.S.","Equipamento","Usuário","Motivo"])
        for r in rows:
            ws.append([r["data_hora"], r["tipo"], r["peca_nome"], r["peca_codigo"],
                       r["quantidade"], r["quantidade_ant"], r["quantidade_pos"],
                       r["os_numero"], r["equipamento_tag"], r["usuario_nome"], r["motivo"]])
        wb.save(caminho)
        messagebox.showinfo("Sucesso", f"Exportado: {caminho}")

    def _rel_criticos(self):
        import openpyxl
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"criticos_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not caminho: return
        rows = db.listar_pecas(apenas_criticos=True)
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Críticos"
        ws.append(["Código","Nome","Categoria","Qtd Atual","Qtd Mínima","Localização","Fornecedor"])
        for r in rows:
            ws.append([r["codigo"], r["nome"], r["categoria_nome"],
                       r["quantidade"], r["estoque_minimo"],
                       r["localizacao_nome"], r["fornecedor_nome"]])
        wb.save(caminho)
        messagebox.showinfo("Sucesso", f"Exportado: {caminho}")

    def _grafico_consumo(self):
        try: import matplotlib.pyplot as plt
        except ImportError:
            messagebox.showerror("Erro", "matplotlib não instalado."); return
        rows = db.get_consumo_mensal()
        if not rows:
            messagebox.showinfo("Info", "Sem dados de movimentação para exibir."); return
        meses  = [r["mes"] for r in rows]
        saidas = [r["saidas"] or 0 for r in rows]
        entradas=[r["entradas"] or 0 for r in rows]
        x = range(len(meses))
        plt.figure(figsize=(12,5), facecolor="#16213e")
        ax = plt.gca(); ax.set_facecolor("#16213e")
        ax.bar([i-0.2 for i in x], entradas, 0.4, label="Entradas", color="#2E7D32")
        ax.bar([i+0.2 for i in x], saidas,   0.4, label="Saídas",   color="#C62828")
        ax.set_xticks(list(x)); ax.set_xticklabels(meses, rotation=45, color="white")
        ax.tick_params(colors="white"); ax.spines[:].set_color("#444")
        ax.yaxis.label.set_color("white"); ax.xaxis.label.set_color("white")
        plt.title("Consumo Mensal — Últimos 12 meses", color="white", fontsize=14)
        plt.legend(facecolor="#1a1a2e", labelcolor="white")
        plt.tight_layout(); plt.show()

    def _grafico_top(self):
        try: import matplotlib.pyplot as plt
        except ImportError:
            messagebox.showerror("Erro", "matplotlib não instalado."); return
        rows = db.get_top_consumo(10)
        if not rows:
            messagebox.showinfo("Info", "Sem dados de consumo."); return
        nomes   = [r["nome"][:25] for r in rows]
        valores = [r["total_saido"] for r in rows]
        plt.figure(figsize=(10,5), facecolor="#16213e")
        ax = plt.gca(); ax.set_facecolor("#16213e")
        bars = ax.barh(nomes[::-1], valores[::-1], color="#1E88E5")
        ax.tick_params(colors="white"); ax.spines[:].set_color("#444")
        plt.title("Top 10 Peças Mais Consumidas (30 dias)", color="white", fontsize=13)
        for bar, val in zip(bars, valores[::-1]):
            ax.text(bar.get_width()+0.1, bar.get_y()+bar.get_height()/2,
                    f"{val:.1f}", va="center", color="white")
        plt.tight_layout(); plt.show()

    def atualizar(self): pass


# ══════════════════════════════════════════════════════════════════════════════
#  PÁGINA: ALERTAS
# ══════════════════════════════════════════════════════════════════════════════

class PaginaAlertas(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=CINZA_BG, corner_radius=0)
        self.app = app
        self._build()

    def _build(self):
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(15,5))
        ctk.CTkLabel(hdr, text="🚨 Alertas de Estoque",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(side="left")
        ctk.CTkButton(hdr, text="✅ Marcar todos como lidos", fg_color=VERDE,
                      command=self._marcar_lidos).pack(side="right")

        self.tabela = Tabela(
            self,
            colunas=["Peça","Código","Qtd Atual","Status","Data","Mensagem"],
            larguras=[200,90,80,100,130,300],
            altura=25
        )
        self.tabela.pack(fill="both", expand=True, padx=20, pady=10)

    def atualizar(self):
        self.tabela.limpar()
        rows = db.get_alertas_nao_lidos()
        for r in rows:
            qtd = r["quantidade"] or 0
            status = "🔴 ZERADO" if qtd == 0 else "⚠️ CRÍTICO"
            tag = "zero" if qtd == 0 else "critico"
            data = r["criado_em"][:16] if r["criado_em"] else ""
            self.tabela.inserir([
                r["peca_nome"], r["peca_codigo"], f"{qtd:.1f}", status, data, r["mensagem"]
            ], tag)
        self.tabela.configurar_tag("zero",    bg="#5D1A1A")
        self.tabela.configurar_tag("critico", bg="#5D4000")

    def _marcar_lidos(self):
        db.marcar_alertas_lidos()
        self.atualizar()
        messagebox.showinfo("OK", "Todos os alertas foram marcados como lidos.")


# ══════════════════════════════════════════════════════════════════════════════
#  PÁGINAS SIMPLES: Equipamentos / Fornecedores / Usuários
# ══════════════════════════════════════════════════════════════════════════════

class PaginaEquipamentos(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=CINZA_BG, corner_radius=0)
        self.app = app
        self._build()

    def _build(self):
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(15,5))
        ctk.CTkLabel(hdr, text="🏭 Equipamentos",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(side="left")
        ctk.CTkButton(hdr, text="＋ Novo Equipamento", fg_color=VERDE,
                      command=self._novo).pack(side="right")

        self.tabela = Tabela(self, colunas=["TAG","Nome","Setor","Modelo","Fabricante"],
                             larguras=[100,200,150,150,150], altura=25)
        self.tabela.pack(fill="both", expand=True, padx=20, pady=10)

    def atualizar(self):
        self.tabela.limpar()
        for e in db.listar_equipamentos():
            self.tabela.inserir([e["tag"], e["nome"], e["setor"] or "",
                                  e["modelo"] or "", e["fabricante"] or ""])

    def _novo(self):
        DialogSimples(self, "Novo Equipamento",
                      campos=["TAG *","Nome *","Setor","Modelo","Fabricante"],
                      callback=self._salvar)

    def _salvar(self, vals):
        tag, nome, setor, modelo, fab = vals
        if not tag or not nome:
            messagebox.showerror("Erro", "TAG e Nome são obrigatórios."); return
        db.inserir_equipamento(tag, nome, setor, modelo, fab)
        self.atualizar()


class PaginaFornecedores(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=CINZA_BG, corner_radius=0)
        self.app = app
        self._build()

    def _build(self):
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(15,5))
        ctk.CTkLabel(hdr, text="🤝 Fornecedores",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(side="left")
        ctk.CTkButton(hdr, text="＋ Novo Fornecedor", fg_color=VERDE,
                      command=self._novo).pack(side="right")

        self.tabela = Tabela(self, colunas=["Nome","CNPJ","Telefone","E-mail","Contato"],
                             larguras=[200,130,120,200,150], altura=25)
        self.tabela.pack(fill="both", expand=True, padx=20, pady=10)

    def atualizar(self):
        self.tabela.limpar()
        for f in db.listar_fornecedores():
            self.tabela.inserir([f["nome"], f["cnpj"] or "", f["telefone"] or "",
                                  f["email"] or "", f["contato"] or ""])

    def _novo(self):
        DialogSimples(self, "Novo Fornecedor",
                      campos=["Nome *","CNPJ","Telefone","E-mail","Contato"],
                      callback=self._salvar)

    def _salvar(self, vals):
        nome, cnpj, tel, email, contato = vals
        if not nome:
            messagebox.showerror("Erro", "Nome é obrigatório."); return
        db.inserir_fornecedor({"nome":nome,"cnpj":cnpj,"telefone":tel,"email":email,"contato":contato})
        self.atualizar()


class PaginaUsuarios(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master, fg_color=CINZA_BG, corner_radius=0)
        self.app = app
        self._build()

    def _build(self):
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(15,5))
        ctk.CTkLabel(hdr, text="👤 Usuários",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(side="left")
        if USUARIO_ATUAL["perfil"] == "admin":
            ctk.CTkButton(hdr, text="＋ Novo Usuário", fg_color=VERDE,
                          command=self._novo).pack(side="right")

        self.tabela = Tabela(self, colunas=["ID","Nome","Login","Perfil","Ativo"],
                             larguras=[50,200,130,100,60], altura=20)
        self.tabela.pack(fill="both", expand=True, padx=20, pady=10)

    def atualizar(self):
        self.tabela.limpar()
        for u in db.listar_usuarios():
            self.tabela.inserir([u["id"], u["nome"], u["login"],
                                  u["perfil"], "Sim" if u["ativo"] else "Não"])

    def _novo(self):
        DialogSimples(self, "Novo Usuário",
                      campos=["Nome *","Login *","Senha *","Perfil (admin/tecnico/consulta)"],
                      callback=self._salvar)

    def _salvar(self, vals):
        nome, login, senha, perfil = vals
        if not nome or not login or not senha:
            messagebox.showerror("Erro", "Nome, Login e Senha são obrigatórios."); return
        perfil = perfil.strip() or "tecnico"
        db.inserir_usuario(nome, login, senha, perfil)
        self.atualizar()
        messagebox.showinfo("Sucesso", f"Usuário '{nome}' criado.")


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITÁRIOS: Diálogos genéricos / Histórico / Etiquetas
# ══════════════════════════════════════════════════════════════════════════════

class DialogSimples(ctk.CTkToplevel):
    def __init__(self, master, titulo, campos, callback):
        super().__init__(master)
        self.title(titulo)
        self.geometry("400x" + str(80 + len(campos) * 70))
        self.resizable(False, False)
        self.configure(fg_color=CINZA_BG)
        self.grab_set()
        self.callback = callback
        self.entries = []

        frame = ctk.CTkFrame(self, fg_color=CINZA_CARD, corner_radius=12)
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(frame, text=titulo, font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(15,10))

        for campo in campos:
            ctk.CTkLabel(frame, text=campo, anchor="w").pack(fill="x", padx=15, pady=(5,2))
            e = ctk.CTkEntry(frame, height=36, show="*" if "Senha" in campo else "")
            e.pack(fill="x", padx=15)
            self.entries.append(e)

        btns = ctk.CTkFrame(frame, fg_color="transparent")
        btns.pack(fill="x", padx=15, pady=15)
        ctk.CTkButton(btns, text="💾 Salvar", fg_color=VERDE,
                      command=self._salvar).pack(side="left", fill="x", expand=True, padx=5)
        ctk.CTkButton(btns, text="✖ Cancelar", fg_color="gray40",
                      command=self.destroy).pack(side="left", fill="x", expand=True, padx=5)

    def _salvar(self):
        vals = [e.get().strip() for e in self.entries]
        self.callback(vals)
        self.destroy()


class HistoricoDialog(ctk.CTkToplevel):
    def __init__(self, master, peca):
        super().__init__(master)
        self.title(f"Histórico — {peca['nome']}")
        self.geometry("900x500")
        self.configure(fg_color=CINZA_BG)
        self.grab_set()

        ctk.CTkLabel(self, text=f"📋 Histórico de Movimentações — {peca['nome']}",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=10)

        tab = Tabela(self,
                     colunas=["Data/Hora","Tipo","Qtd","Ant.","Pos.","Equipamento","O.S.","Usuário","Motivo"],
                     larguras=[130,100,60,60,60,130,80,100,200], altura=18)
        tab.pack(fill="both", expand=True, padx=15, pady=(0,15))

        rows = db.listar_movimentacoes(peca_id=peca["id"])
        for r in rows:
            data = r["data_hora"][:16] if r["data_hora"] else ""
            tab.inserir([
                data, label_tipo(r["tipo"]),
                f"{r['quantidade']:.1f}", f"{r['quantidade_ant']:.1f}", f"{r['quantidade_pos']:.1f}",
                r["equipamento_tag"] or "", r["os_numero"] or "",
                r["usuario_nome"] or "", r["motivo"] or ""
            ], r["tipo"])
        tab.configurar_tag("entrada",   bg="#1B5E20")
        tab.configurar_tag("saida",     bg="#7f0000")
        tab.configurar_tag("ajuste",    bg="#33691E")
        tab.configurar_tag("devolucao", bg="#1A237E")


def gerar_etiqueta(peca):
    """Gera etiqueta com código de barras Code128 e QR Code e salva em PNG."""
    try:
        import barcode as pybr
        from barcode.writer import ImageWriter
        import qrcode
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        messagebox.showerror("Erro", "Instale: pip install python-barcode qrcode pillow")
        return

    caminho = filedialog.asksaveasfilename(
        defaultextension=".png", filetypes=[("PNG","*.png")],
        initialfile=f"etiqueta_{peca['codigo']}.png"
    )
    if not caminho: return

    # Dimensões da etiqueta (400 x 220 px)
    W, H = 400, 220
    img = Image.new("RGB", (W, H), "#ffffff")
    draw = ImageDraw.Draw(img)

    # Cabeçalho
    draw.rectangle([0,0,W,40], fill="#1565C0")
    try:
        font_titulo = ImageFont.truetype("arial.ttf", 14)
        font_nome   = ImageFont.truetype("arial.ttf", 12)
        font_small  = ImageFont.truetype("arial.ttf", 10)
    except:
        font_titulo = ImageFont.load_default()
        font_nome   = font_titulo
        font_small  = font_titulo

    draw.text((10, 12), "⚙️ CONTROLE DE ESTOQUE — MANUTENÇÃO", fill="white", font=font_titulo)

    # Nome e código
    draw.text((10, 48), peca["nome"][:40], fill="#000000", font=font_nome)
    draw.text((10, 68), f"Cód: {peca['codigo']}", fill="#333333", font=font_small)
    cat = peca.get("categoria_nome") or ""
    loc = peca.get("localizacao_nome") or ""
    draw.text((10, 84), f"Cat: {cat}  |  Local: {loc}", fill="#555555", font=font_small)
    draw.text((10, 100), f"Qtd: {peca['quantidade']:.1f} {peca['unidade']}  |  Mín: {peca['estoque_minimo']:.1f}", fill="#555555", font=font_small)

    # Código de barras
    barcode_val = peca.get("codigo_barras") or peca["codigo"]
    try:
        cls = pybr.get_barcode_class("code128")
        buf = io.BytesIO()
        cls(barcode_val, writer=ImageWriter()).write(buf,
            options={"module_width":0.8, "module_height":8, "write_text":True})
        buf.seek(0)
        bar_img = Image.open(buf).resize((230, 80))
        img.paste(bar_img, (10, 120))
    except Exception:
        draw.text((10,130), f"[ {barcode_val} ]", fill="black", font=font_nome)

    # QR Code
    qr = qrcode.QRCode(box_size=3, border=2)
    qr.add_data(f"COD:{peca['codigo']}|NOME:{peca['nome']}|QTD:{peca['quantidade']}")
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").resize((80, 80))
    img.paste(qr_img, (310, 120))

    img.save(caminho)
    messagebox.showinfo("Etiqueta gerada!", f"Etiqueta salva em:\n{caminho}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    db.inicializar_banco()

    # Janela raiz oculta (necessária para o Toplevel de login)
    root = ctk.CTk()
    root.withdraw()

    def abrir_app():
        root.destroy()
        app = App()
        app.mainloop()

    login = TelaLogin(abrir_app)
    login.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()


if __name__ == "__main__":
    main()
