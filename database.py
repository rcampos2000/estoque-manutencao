"""
database.py — Banco de dados SQLite para o sistema de controle de peças de reposição
"""

import sqlite3
import hashlib
import os
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "estoque_manutencao.db")


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode()).hexdigest()


def inicializar_banco():
    conn = get_connection()
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS categorias (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            nome      TEXT NOT NULL UNIQUE,
            descricao TEXT,
            cor       TEXT DEFAULT '#3498db',
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS fornecedores (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            nome      TEXT NOT NULL,
            cnpj      TEXT,
            telefone  TEXT,
            email     TEXT,
            contato   TEXT,
            ativo     INTEGER DEFAULT 1,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS localizacoes (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            nome      TEXT NOT NULL UNIQUE,
            descricao TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS equipamentos (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            tag        TEXT NOT NULL UNIQUE,
            nome       TEXT NOT NULL,
            setor      TEXT,
            modelo     TEXT,
            fabricante TEXT,
            ativo      INTEGER DEFAULT 1,
            criado_em  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS pecas (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo           TEXT NOT NULL UNIQUE,
            codigo_barras    TEXT UNIQUE,
            nome             TEXT NOT NULL,
            descricao        TEXT,
            categoria_id     INTEGER REFERENCES categorias(id),
            fornecedor_id    INTEGER REFERENCES fornecedores(id),
            localizacao_id   INTEGER REFERENCES localizacoes(id),
            unidade          TEXT DEFAULT 'UN',
            quantidade       REAL DEFAULT 0,
            estoque_minimo   REAL DEFAULT 1,
            estoque_maximo   REAL DEFAULT 100,
            custo_unitario   REAL DEFAULT 0,
            preco_venda      REAL DEFAULT 0,
            imagem_path      TEXT,
            ativo            INTEGER DEFAULT 1,
            criado_em        TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            atualizado_em    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            nome      TEXT NOT NULL,
            login     TEXT NOT NULL UNIQUE,
            senha     TEXT NOT NULL,
            perfil    TEXT DEFAULT 'tecnico',
            ativo     INTEGER DEFAULT 1,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS movimentacoes (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            peca_id        INTEGER NOT NULL REFERENCES pecas(id),
            tipo           TEXT NOT NULL,
            quantidade     REAL NOT NULL,
            quantidade_ant REAL NOT NULL DEFAULT 0,
            quantidade_pos REAL NOT NULL DEFAULT 0,
            custo_unitario REAL DEFAULT 0,
            usuario_id     INTEGER REFERENCES usuarios(id),
            equipamento_id INTEGER REFERENCES equipamentos(id),
            os_numero      TEXT,
            motivo         TEXT,
            observacao     TEXT,
            data_hora      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS alertas (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            peca_id   INTEGER REFERENCES pecas(id),
            tipo      TEXT,
            mensagem  TEXT,
            lido      INTEGER DEFAULT 0,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Admin padrão
    c.execute("SELECT id FROM usuarios WHERE login='admin'")
    if not c.fetchone():
        c.execute("INSERT INTO usuarios (nome, login, senha, perfil) VALUES (?,?,?,?)",
                  ("Administrador", "admin", hash_senha("admin123"), "admin"))

    # Categorias padrão
    cats = [
        ("Rolamentos",    "Rolamentos e mancais",       "#e74c3c"),
        ("Correias",      "Correias e transmissão",      "#e67e22"),
        ("Elétricos",     "Componentes elétricos",       "#f1c40f"),
        ("Hidráulicos",   "Sistema hidráulico",          "#3498db"),
        ("Pneumáticos",   "Sistema pneumático",          "#2ecc71"),
        ("Vedações",      "Retentores, O-rings, gaxetas","#9b59b6"),
        ("Filtros",       "Filtros em geral",            "#1abc9c"),
        ("Lubrificantes", "Óleos e graxas",              "#34495e"),
        ("Outros",        "Itens diversos",              "#95a5a6"),
    ]
    for nome, desc, cor in cats:
        c.execute("INSERT OR IGNORE INTO categorias (nome, descricao, cor) VALUES (?,?,?)",
                  (nome, desc, cor))

    locs = ["Almoxarifado A", "Almoxarifado B", "Gaveta Elétrica", "Prateleira 1", "Prateleira 2"]
    for loc in locs:
        c.execute("INSERT OR IGNORE INTO localizacoes (nome) VALUES (?)", (loc,))

    conn.commit()
    conn.close()


# ── CRUD Peças ────────────────────────────────────────────────────────────────

def listar_pecas(filtro="", categoria_id=None, apenas_criticos=False):
    conn = get_connection()
    c = conn.cursor()
    q = """
        SELECT p.*, cat.nome AS categoria_nome, cat.cor AS categoria_cor,
               loc.nome AS localizacao_nome, forn.nome AS fornecedor_nome
        FROM pecas p
        LEFT JOIN categorias cat ON p.categoria_id = cat.id
        LEFT JOIN localizacoes loc ON p.localizacao_id = loc.id
        LEFT JOIN fornecedores forn ON p.fornecedor_id = forn.id
        WHERE p.ativo = 1
    """
    params = []
    if filtro:
        q += " AND (p.nome LIKE ? OR p.codigo LIKE ? OR p.codigo_barras LIKE ? OR p.descricao LIKE ?)"
        like = f"%{filtro}%"
        params.extend([like, like, like, like])
    if categoria_id:
        q += " AND p.categoria_id = ?"
        params.append(categoria_id)
    if apenas_criticos:
        q += " AND p.quantidade <= p.estoque_minimo"
    q += " ORDER BY p.nome"
    c.execute(q, params)
    rows = c.fetchall()
    conn.close()
    return rows


def buscar_peca_barcode(codigo_barras):
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT p.*, cat.nome AS categoria_nome, loc.nome AS localizacao_nome, forn.nome AS fornecedor_nome
        FROM pecas p
        LEFT JOIN categorias cat ON p.categoria_id = cat.id
        LEFT JOIN localizacoes loc ON p.localizacao_id = loc.id
        LEFT JOIN fornecedores forn ON p.fornecedor_id = forn.id
        WHERE p.codigo_barras = ? OR p.codigo = ?
    """, (codigo_barras, codigo_barras))
    row = c.fetchone()
    conn.close()
    return row


def inserir_peca(dados: dict):
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        INSERT INTO pecas (codigo, codigo_barras, nome, descricao, categoria_id, fornecedor_id,
                           localizacao_id, unidade, quantidade, estoque_minimo, estoque_maximo,
                           custo_unitario, preco_venda)
        VALUES (:codigo, :codigo_barras, :nome, :descricao, :categoria_id, :fornecedor_id,
                :localizacao_id, :unidade, :quantidade, :estoque_minimo, :estoque_maximo,
                :custo_unitario, :preco_venda)
    """, dados)
    conn.commit()
    peca_id = c.lastrowid
    conn.close()
    return peca_id


def atualizar_peca(peca_id, dados: dict):
    conn = get_connection()
    c = conn.cursor()
    dados["id"] = peca_id
    dados["atualizado_em"] = datetime.now().isoformat()
    c.execute("""
        UPDATE pecas SET
            codigo=:codigo, codigo_barras=:codigo_barras, nome=:nome, descricao=:descricao,
            categoria_id=:categoria_id, fornecedor_id=:fornecedor_id,
            localizacao_id=:localizacao_id, unidade=:unidade,
            estoque_minimo=:estoque_minimo, estoque_maximo=:estoque_maximo,
            custo_unitario=:custo_unitario, preco_venda=:preco_venda,
            atualizado_em=:atualizado_em
        WHERE id=:id
    """, dados)
    conn.commit()
    conn.close()


def excluir_peca(peca_id):
    conn = get_connection()
    conn.execute("UPDATE pecas SET ativo=0 WHERE id=?", (peca_id,))
    conn.commit()
    conn.close()


# ── Movimentações ─────────────────────────────────────────────────────────────

def registrar_movimentacao(peca_id, tipo, quantidade, usuario_id,
                            custo=0, equip_id=None, os_num=None, motivo=None, obs=None):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT quantidade FROM pecas WHERE id=?", (peca_id,))
    row = c.fetchone()
    qtd_ant = row["quantidade"] if row else 0

    if tipo in ("entrada", "devolucao"):
        qtd_pos = qtd_ant + quantidade
    elif tipo == "saida":
        qtd_pos = qtd_ant - quantidade
    elif tipo == "ajuste":
        qtd_pos = quantidade
        quantidade = abs(quantidade - qtd_ant)
    else:
        qtd_pos = qtd_ant

    c.execute("""
        INSERT INTO movimentacoes
            (peca_id, tipo, quantidade, quantidade_ant, quantidade_pos,
             custo_unitario, usuario_id, equipamento_id, os_numero, motivo, observacao)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (peca_id, tipo, abs(quantidade), qtd_ant, qtd_pos,
          custo, usuario_id, equip_id, os_num, motivo, obs))

    c.execute("UPDATE pecas SET quantidade=?, atualizado_em=? WHERE id=?",
              (qtd_pos, datetime.now().isoformat(), peca_id))

    c.execute("SELECT nome, estoque_minimo FROM pecas WHERE id=?", (peca_id,))
    p = c.fetchone()
    if p and qtd_pos <= p["estoque_minimo"]:
        msg = f"Estoque crítico: {p['nome']} — apenas {qtd_pos:.1f} em estoque"
        c.execute("INSERT INTO alertas (peca_id, tipo, mensagem) VALUES (?, 'estoque_minimo', ?)",
                  (peca_id, msg))

    conn.commit()
    conn.close()
    return qtd_pos


def listar_movimentacoes(peca_id=None, tipo=None, data_inicio=None, data_fim=None, limit=300):
    conn = get_connection()
    c = conn.cursor()
    q = """
        SELECT m.*, p.nome AS peca_nome, p.codigo AS peca_codigo,
               u.nome AS usuario_nome, e.tag AS equipamento_tag, e.nome AS equipamento_nome
        FROM movimentacoes m
        LEFT JOIN pecas p ON m.peca_id = p.id
        LEFT JOIN usuarios u ON m.usuario_id = u.id
        LEFT JOIN equipamentos e ON m.equipamento_id = e.id
        WHERE 1=1
    """
    params = []
    if peca_id:
        q += " AND m.peca_id=?"; params.append(peca_id)
    if tipo:
        q += " AND m.tipo=?"; params.append(tipo)
    if data_inicio:
        q += " AND DATE(m.data_hora) >= ?"; params.append(data_inicio)
    if data_fim:
        q += " AND DATE(m.data_hora) <= ?"; params.append(data_fim)
    q += f" ORDER BY m.data_hora DESC LIMIT {limit}"
    c.execute(q, params)
    rows = c.fetchall()
    conn.close()
    return rows


# ── Usuários ──────────────────────────────────────────────────────────────────

def autenticar(login, senha):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM usuarios WHERE login=? AND senha=? AND ativo=1",
              (login, hash_senha(senha)))
    user = c.fetchone()
    conn.close()
    return user


def listar_usuarios():
    conn = get_connection()
    rows = conn.execute("SELECT id, nome, login, perfil, ativo FROM usuarios ORDER BY nome").fetchall()
    conn.close()
    return rows


def inserir_usuario(nome, login, senha, perfil="tecnico"):
    conn = get_connection()
    conn.execute("INSERT INTO usuarios (nome, login, senha, perfil) VALUES (?,?,?,?)",
                 (nome, login, hash_senha(senha), perfil))
    conn.commit()
    conn.close()


def alterar_senha(usuario_id, nova_senha):
    conn = get_connection()
    conn.execute("UPDATE usuarios SET senha=? WHERE id=?",
                 (hash_senha(nova_senha), usuario_id))
    conn.commit()
    conn.close()


# ── Auxiliares ────────────────────────────────────────────────────────────────

def listar_categorias():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM categorias ORDER BY nome").fetchall()
    conn.close()
    return rows


def listar_fornecedores(apenas_ativos=True):
    conn = get_connection()
    q = "SELECT * FROM fornecedores" + (" WHERE ativo=1" if apenas_ativos else "") + " ORDER BY nome"
    rows = conn.execute(q).fetchall()
    conn.close()
    return rows


def inserir_fornecedor(dados):
    conn = get_connection()
    conn.execute("INSERT INTO fornecedores (nome, cnpj, telefone, email, contato) VALUES (?,?,?,?,?)",
                 (dados.get("nome"), dados.get("cnpj"), dados.get("telefone"),
                  dados.get("email"), dados.get("contato")))
    conn.commit()
    conn.close()


def listar_localizacoes():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM localizacoes ORDER BY nome").fetchall()
    conn.close()
    return rows


def listar_equipamentos():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM equipamentos WHERE ativo=1 ORDER BY tag").fetchall()
    conn.close()
    return rows


def inserir_equipamento(tag, nome, setor="", modelo="", fabricante=""):
    conn = get_connection()
    conn.execute("INSERT INTO equipamentos (tag, nome, setor, modelo, fabricante) VALUES (?,?,?,?,?)",
                 (tag, nome, setor, modelo, fabricante))
    conn.commit()
    conn.close()


# ── Dashboard / KPIs ──────────────────────────────────────────────────────────

def get_kpis():
    conn = get_connection()
    c = conn.cursor()
    kpis = {}
    c.execute("SELECT COUNT(*) FROM pecas WHERE ativo=1"); kpis["total_pecas"] = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM pecas WHERE ativo=1 AND quantidade <= estoque_minimo")
    kpis["criticos"] = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM pecas WHERE ativo=1 AND quantidade = 0")
    kpis["zerados"] = c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(quantidade * custo_unitario),0) FROM pecas WHERE ativo=1")
    kpis["valor_total"] = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM movimentacoes WHERE tipo='saida' AND DATE(data_hora) >= DATE('now','-30 days')")
    kpis["saidas_30d"] = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM movimentacoes WHERE tipo='entrada' AND DATE(data_hora) >= DATE('now','-30 days')")
    kpis["entradas_30d"] = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM alertas WHERE lido=0"); kpis["alertas"] = c.fetchone()[0]
    conn.close()
    return kpis


def get_consumo_mensal():
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT strftime('%Y-%m', data_hora) AS mes,
               SUM(CASE WHEN tipo='saida'   THEN quantidade ELSE 0 END) AS saidas,
               SUM(CASE WHEN tipo='entrada' THEN quantidade ELSE 0 END) AS entradas
        FROM movimentacoes
        WHERE data_hora >= DATE('now', '-12 months')
        GROUP BY mes ORDER BY mes
    """)
    rows = c.fetchall()
    conn.close()
    return rows


def get_top_consumo(limit=10):
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT p.nome, p.codigo, SUM(m.quantidade) AS total_saido
        FROM movimentacoes m
        JOIN pecas p ON m.peca_id = p.id
        WHERE m.tipo='saida' AND m.data_hora >= DATE('now','-30 days')
        GROUP BY m.peca_id ORDER BY total_saido DESC LIMIT ?
    """, (limit,))
    rows = c.fetchall()
    conn.close()
    return rows


def get_alertas_nao_lidos():
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT a.*, p.nome AS peca_nome, p.codigo AS peca_codigo, p.quantidade
        FROM alertas a JOIN pecas p ON a.peca_id = p.id
        WHERE a.lido=0 ORDER BY a.criado_em DESC
    """)
    rows = c.fetchall()
    conn.close()
    return rows


def marcar_alertas_lidos():
    conn = get_connection()
    conn.execute("UPDATE alertas SET lido=1")
    conn.commit()
    conn.close()


def exportar_estoque_excel(caminho):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT p.codigo, p.codigo_barras, p.nome, p.descricao,
               cat.nome AS categoria, p.unidade, p.quantidade,
               p.estoque_minimo, p.estoque_maximo,
               p.custo_unitario, (p.quantidade * p.custo_unitario) AS valor_total,
               loc.nome AS localizacao, forn.nome AS fornecedor
        FROM pecas p
        LEFT JOIN categorias cat ON p.categoria_id = cat.id
        LEFT JOIN localizacoes loc ON p.localizacao_id = loc.id
        LEFT JOIN fornecedores forn ON p.fornecedor_id = forn.id
        WHERE p.ativo=1 ORDER BY cat.nome, p.nome
    """)
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque"

    headers = ["Código", "Cód.Barras", "Nome", "Descrição", "Categoria", "Unid.",
               "Quantidade", "Est.Mín.", "Est.Máx.", "Custo Unit.", "Valor Total",
               "Localização", "Fornecedor"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    red_fill   = PatternFill("solid", fgColor="FFCCCC")
    green_fill = PatternFill("solid", fgColor="CCFFCC")

    for r, row in enumerate(rows, 2):
        vals = list(row)
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        qtd = row["quantidade"] if row["quantidade"] else 0
        est_min = row["estoque_minimo"] if row["estoque_minimo"] else 0
        fill = red_fill if qtd <= est_min else (green_fill if r % 2 == 0 else None)
        if fill:
            for col in range(1, len(headers)+1):
                ws.cell(row=r, column=col).fill = fill

    col_widths = [12, 14, 35, 30, 15, 6, 10, 8, 8, 12, 12, 18, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(caminho)
    return caminho
